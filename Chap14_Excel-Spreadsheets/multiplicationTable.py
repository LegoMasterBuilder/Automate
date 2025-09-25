# multiplicationTable.py

# Create a program multiplicationTable.py that takes a number N 
# from the command line and creates an N×N multiplication table 
# in an Excel spreadsheet. For example, when the program is run 
# like this: 
# py multiplicationTable.py 6

import openpyxl, sys
from openpyxl.utils import get_column_letter

try:
    # Obtain value for n from user input.
    n = int(sys.argv[1])

    # Generate new workbook.
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Iterate through the workbook cells.
    for row in range(0, n+1):
        for col in range(0, n+1):
            # Get the cell coordinates.
            cell = str(get_column_letter(row+1)) + str(col+1)

            # Check for header cells.
            if row == 0 or col == 0:
                if row == 0 and col == 0:
                    continue
                elif row == 0:
                    # Supply y-axis headers
                    sheet[cell] = col
                elif col == 0:
                    # Supply x-axis headers
                    sheet[cell] = row
            else:
                # print(f'{cell}, {row*col}')
                sheet[cell] = row * col

    # Save changes to a .xlsx file.
    wb.save('multiplicationTable.xlsx')

# Error handling for incomplete user input.
except IndexError:
    print("Please try again. Enter value for n.")



