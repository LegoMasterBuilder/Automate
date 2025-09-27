#blankRowInserter.py

# Create a program blankRowInserter.py that takes two integers and a filename 
# string as command line arguments. Let’s call the first integer N and the 
# second integer M. Starting at row N, the program should insert M blank rows 
# into the spreadsheet. For example, when the program is run like this:

# python blankRowInserter.py 3 2 myProduce.xlsx

import sys, openpyxl
from openpyxl.utils import get_column_letter

try:
    # Set row N and M blank rows.
    N, M = int(sys.argv[1]), int(sys.argv[2])
    spreadsheet = sys.argv[3]

    # Open existing workbook
    wb1 = openpyxl.load_workbook(spreadsheet)
    sheetOrig = wb1['Sheet1']

    # Generate new workbook.
    wb2 = openpyxl.Workbook()
    sheetNew = wb2['Sheet']

    # Iterate through workbook.
    for row in range(1, sheetOrig.max_row + M + 1):
        for col in range(1, sheetOrig.max_column + 1):
            # Extract Column and Row Numbers.
            column_letter = str(get_column_letter(col))
            row_number_old = str(row)
            row_number_new = str(row + M)

            # Extract cell locations.
            cell_loc_old = column_letter + row_number_old
            cell_loc_new = column_letter + row_number_new

            while row in range(N, N+M):
                sheetNew[cell_loc_new] = " "
            else: 
                sheetNew[cell_loc_new] = sheetOrig[cell_loc_old].value
            
    # Save changes to a .xlsx file.
    wb2.save('editBlank.xlsx')

except IndexError:
    print("Incomplete input. Please try again.")

except FileNotFoundError:
    print("File input not found. Please try again.")